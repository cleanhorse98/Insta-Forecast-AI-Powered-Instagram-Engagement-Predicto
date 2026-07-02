import re
import os
import json
from datetime import datetime
from typing import Dict, Any, List, Optional
import instaloader

# import instagram_scraper  # Removed incorrect self-import
import requests
import random
try:
    import vpn  # optional: used to toggle WARP on rate-limit
except Exception:
    vpn = None

# Utility: extract shortcode from a post/reel URL
def shortcode_from_url(url: str) -> Optional[str]:
    # matches /p/shortcode/ or /reel/shortcode/ or /tv/shortcode/
    m = re.search(r"(?:/p/|/reel/|/tv/)([A-Za-z0-9_-]+)", url)
    return m.group(1) if m else None

# Main extractor using Instaloader
def get_instagram_post_data(url: str, login_user: Optional[str] = None, login_pass: Optional[str] = None,
                            fetch_comments: int = 20,
                            session_file: Optional[str] = None,
                            proxy: Optional[str] = None,
                            max_retries: int = 3,
                            retry_wait: int = 15,
                            min_wait: float = 0.0,
                            max_wait: float = 0.0,
                            requests_per_minute: Optional[int] = None,
                            proxy_pool: Optional[List[str]] = None,
                            user_agent: Optional[str] = None,
                            user_agent_pool: Optional[List[str]] = None) -> Dict[str, Any]:
    """
    Returns a dict with available public metadata for the given Instagram post/reel URL.
    If you provide login_user and login_pass, the script will log in (useful for higher rate limits or private posts you can access).
    Note: 'shares' and account-level insights are NOT available via public Instaloader. Use the Instagram Graph API for insights.
    """
    shortcode = shortcode_from_url(url)
    if not shortcode:
        raise ValueError("Could not parse shortcode from URL. Provide a valid Instagram post/reel URL.")

    L = instaloader.Instaloader()
    # Optional proxy support
    if proxy:
        try:
            L.context._session.proxies.update({
                "http": proxy,
                "https": proxy,
            })
        except Exception:
            pass
    # Optional proxy pool (round-robin per attempt)
    pool_proxies = proxy_pool or []

    # Optional user-agent override/rotation
    try:
        if user_agent:
            L.context._session.headers.update({"User-Agent": user_agent})
        elif user_agent_pool:
            L.context._session.headers.update({"User-Agent": random.choice(user_agent_pool)})
    except Exception:
        pass

    # Session handling: prefer loading an existing session if provided
    if session_file:
        try:
            L.load_session_from_file(username=login_user or "", filename=session_file)
        except Exception:
            # If load fails and credentials provided, login and save
            if login_user and login_pass:
                try:
                    L.login(login_user, login_pass)
                    try:
                        L.save_session_to_file(filename=session_file)
                    except Exception:
                        pass
                except Exception:
                    pass
    elif login_user and login_pass:
        try:
            L.login(login_user, login_pass)
        except Exception:
            pass

    # Rate limiting helpers
    def maybe_sleep_between_requests(multiplier: int = 1) -> None:
        try:
            # simple RPM gate: sleep evenly if configured
            if requests_per_minute and requests_per_minute > 0:
                import time
                time.sleep(max(0.0, 60.0 / float(requests_per_minute)))
            # random jitter wait
            if max_wait and max_wait > 0:
                lo = max(0.0, min_wait or 0.0)
                hi = max(lo, max_wait)
                time.sleep(random.uniform(lo, hi) * max(1, multiplier))
        except Exception:
            pass

    # Fetch post with retry/backoff to handle transient 401/403/429
    attempt = 0
    last_err: Optional[Exception] = None
    while attempt <= max_retries:
        try:
            # rotate proxy and user-agent if pools provided
            if pool_proxies:
                try:
                    sel_proxy = pool_proxies[attempt % len(pool_proxies)]
                    L.context._session.proxies.update({"http": sel_proxy, "https": sel_proxy})
                except Exception:
                    pass
            if user_agent_pool:
                try:
                    L.context._session.headers.update({"User-Agent": random.choice(user_agent_pool)})
                except Exception:
                    pass
            maybe_sleep_between_requests()
            post = instaloader.Post.from_shortcode(L.context, shortcode)
            break
        except Exception as e:
            last_err = e
            attempt += 1
            if attempt > max_retries:
                raise e
            # exponential backoff
            try:
                import time
                # add jittered backoff
                base = max(1, retry_wait)
                jitter = random.uniform(0.5, 1.5)
                maybe_sleep_between_requests(multiplier=attempt)
                time.sleep(base * attempt * jitter)
            except Exception:
                pass

    # caption and hashtags
    caption = post.caption or ""
    hashtags = list(post.caption_hashtags) if hasattr(post, "caption_hashtags") else re.findall(r"#\w+", caption)

    # media URLs: handle sidecar (multiple media) or single
    media_urls: List[str] = []
    try:
        if post.is_sidecar:
            for node in post.get_sidecar_nodes():
                if node.is_video:
                    media_urls.append(node.video_url or node.display_url)
                else:
                    media_urls.append(node.display_url)
        else:
            if post.is_video:
                media_urls.append(getattr(post, "video_url", post.url))
            else:
                media_urls.append(post.url)
    except Exception:
        # fallback
        media_urls = [post.url]

    # basic metrics
    likes = post.likes if hasattr(post, "likes") else None
    comments_count = post.comments if hasattr(post, "comments") else None
    timestamp = post.date_utc.isoformat() if hasattr(post, "date_utc") else None
    is_video = getattr(post, "is_video", False)
    video_view_count = getattr(post, "video_view_count", None)  # may be None for photos
    owner_username = getattr(post, "owner_username", None)
    owner_id = getattr(post, "owner_id", None)
    location = None
    if getattr(post, "location", None):
        location = {
            "name": post.location.name,
            "lat": getattr(post.location, "lat", None),
            "lng": getattr(post.location, "lng", None),
        }

    # fetch first N comments (text + owner) with throttling
    comments: List[Dict[str, Any]] = []
    try:
        for i, c in enumerate(post.get_comments()):
            if i >= fetch_comments:
                break
            maybe_sleep_between_requests()
            comments.append({
                "id": getattr(c, "id", None),
                "owner_username": getattr(c, "owner", None).username if getattr(c, "owner", None) else None,
                "text": getattr(c, "text", None),
                "created_at_utc": getattr(c, "created_at_utc", None).isoformat() if getattr(c, "created_at_utc", None) else None
            })
    except Exception:
        # if comments cannot be iterated due to rate limits or privacy, ignore
        comments = []

    result = {
        "shortcode": shortcode,
        "url": url,
        "caption": caption,
        "hashtags": hashtags,
        "likes": likes,
        "comments_count": comments_count,
        "comments_sample": comments,
        "is_video": is_video,
        "video_view_count": video_view_count,
        "timestamp_utc": timestamp,
        "media_urls": media_urls,
        "owner_username": owner_username,
        "owner_id": owner_id,
        "location": location,
        # Not available via Instaloader / public scraping:
        "shares_count": None,
        "insights": None,
    }

    return result

# Helper: call Instagram Graph API for insights (requires Business/Creator account + access token)
def fetch_graph_api_insights(ig_media_id: str, access_token: str, metrics: Optional[List[str]] = None) -> Dict[str, Any]:
    """
    Example usage:
      fetch_graph_api_insights('<IG_MEDIA_ID>', '<ACCESS_TOKEN>', metrics=['impressions','reach','engagement','saved','video_views'])
    Notes:
      - ig_media_id is the Facebook/IG media id (numeric) for the Instagram media object.
      - Access token must have instagram_basic and instagram_manage_insights (and the IG account must be Business/Creator).
      - Graph API version should be adjusted as needed.
    """
    if metrics is None:
        metrics = ["impressions", "reach", "engagement", "saved", "video_views"]

    endpoint = f"https://graph.facebook.com/v17.0/{ig_media_id}/insights"
    params = {
        "metric": ",".join(metrics),
        "access_token": access_token,
    }
    resp = requests.get(endpoint, params=params, timeout=20)
    resp.raise_for_status()
    return resp.json()

def compute_seo_features(data: Dict[str, Any]) -> Dict[str, Any]:
    """
    Compute lightweight SEO features and a score (0-100) from the post data.
    Uses caption length, hashtag count, mentions, URLs, likes/comments, recency and media count.
    """
    caption = data.get("caption", "") or ""
    hashtags = data.get("hashtags", []) or []
    likes = data.get("likes") or 0
    comments_count = data.get("comments_count") or 0
    media_count = len(data.get("media_urls", []) or [])
    timestamp = data.get("timestamp_utc")

    # basic counts
    caption_len = len(caption)
    word_count = len(caption.split())
    hashtag_count = len(hashtags)
    mention_count = len(re.findall(r"@\w+", caption))
    url_count = len(re.findall(r"https?://", caption))

    # engagement proxy (normalize by an arbitrary ceiling)
    eng = min(1.0, (likes + comments_count) / max(1.0, 1000.0))

    # recency in days -> normalized (newer = higher)
    recency_norm = 0.0
    if timestamp:
        try:
            dt = datetime.fromisoformat(timestamp)
            days = (datetime.utcnow() - dt).days
            recency_norm = max(0.0, 1.0 - min(days / 365.0, 1.0))
        except Exception:
            recency_norm = 0.0

    # hashtag contribution (saturating)
    hashtag_norm = min(1.0, hashtag_count / 10.0)

    # media richness
    media_norm = min(1.0, media_count / 3.0)

    # weighted score (result 0-100)
    score = 10 + 25 * hashtag_norm + 25 * eng + 20 * recency_norm + 20 * media_norm
    score = max(0.0, min(100.0, score))

    return {
        "caption_len": caption_len,
        "word_count": word_count,
        "hashtag_count": hashtag_count,
        "mention_count": mention_count,
        "url_count": url_count,
        "likes": likes,
        "comments_count": comments_count,
        "media_count": media_count,
        "recency_norm": recency_norm,
        "engagement_norm": eng,
        "hashtag_norm": hashtag_norm,
        "media_norm": media_norm,
        "seo_score": round(score, 2)
    }

def get_full_instagram_post_data(url: str,
                                 login_user: Optional[str] = None,
                                 login_pass: Optional[str] = None,
                                 fetch_comments: int = 20,
                                 access_token: Optional[str] = None,
                                 save_media: bool = False,
                                 media_dir: str = "media",
                                 session_file: Optional[str] = None,
                                 proxy: Optional[str] = None,
                                 max_retries: int = 3,
                                 retry_wait: int = 15,
                                 min_wait: float = 0.0,
                                 max_wait: float = 0.0,
                                 requests_per_minute: Optional[int] = None,
                                 proxy_pool: Optional[List[str]] = None,
                                 user_agent: Optional[str] = None,
                                 user_agent_pool: Optional[List[str]] = None) -> Dict[str, Any]:
    """
    Wrapper that returns the base data from [`get_instagram_post_data`](instaloader.py),
    optionally fetches Graph API insights via [`fetch_graph_api_insights`](instaloader.py) if access_token and ig_media_id exist,
    computes SEO features via [`compute_seo_features`](instaloader.py), and optionally downloads media.
    """
    # reuse existing extractor
    data = get_instagram_post_data(
        url,
        login_user=login_user,
        login_pass=login_pass,
        fetch_comments=fetch_comments,
        session_file=session_file,
        proxy=proxy,
        max_retries=max_retries,
        retry_wait=retry_wait,
        min_wait=min_wait,
        max_wait=max_wait,
        requests_per_minute=requests_per_minute,
        proxy_pool=proxy_pool,
        user_agent=user_agent,
        user_agent_pool=user_agent_pool,
    )

    # try to find a numeric media id if available via instaloader internals (best-effort)
    ig_media_id = None
    try:
        L = instaloader.Instaloader()
        if proxy:
            try:
                L.context._session.proxies.update({
                    "http": proxy,
                    "https": proxy,
                })
            except Exception:
                pass
        # honor UA headers here too
        try:
            if user_agent:
                L.context._session.headers.update({"User-Agent": user_agent})
            elif user_agent_pool:
                L.context._session.headers.update({"User-Agent": random.choice(user_agent_pool)})
        except Exception:
            pass
        if session_file:
            try:
                L.load_session_from_file(username=login_user or "", filename=session_file)
            except Exception:
                if login_user and login_pass:
                    try:
                        L.login(login_user, login_pass)
                        try:
                            L.save_session_to_file(filename=session_file)
                        except Exception:
                            pass
                    except Exception:
                        pass
        elif login_user and login_pass:
            try:
                L.login(login_user, login_pass)
            except Exception:
                pass
        sc = shortcode_from_url(url)
        post = instaloader.Post.from_shortcode(L.context, sc)
        ig_media_id = getattr(post, "mediaid", None) or getattr(post, "media_id", None) or getattr(post, "media_pk", None)
    except Exception:
        ig_media_id = None

    # fetch Graph API insights if requested and we have an id
    insights = None
    if access_token and ig_media_id:
        try:
            insights = fetch_graph_api_insights(str(ig_media_id), access_token)
        except Exception:
            insights = None

    data["insights"] = insights
    data["ig_media_id"] = ig_media_id

    # compute SEO features & score
    seo = compute_seo_features(data)
    data["seo"] = seo

    # optional media download
    if save_media:
        os.makedirs(media_dir, exist_ok=True)
        saved_files = []
        for i, murl in enumerate(data.get("media_urls", []) or []):
            try:
                # throttle media downloads too
                try:
                    import time
                    if requests_per_minute and requests_per_minute > 0:
                        time.sleep(max(0.0, 60.0 / float(requests_per_minute)))
                    if max_wait and max_wait > 0:
                        time.sleep(random.uniform(max(0.0, min_wait or 0.0), max_wait))
                except Exception:
                    pass
                r = requests.get(murl, timeout=30)
                r.raise_for_status()
                ext = os.path.splitext(murl.split("?")[0])[1] or ".jpg"
                fname = f"{data.get('shortcode', 'post')}_{i}{ext}"
                path = os.path.join(media_dir, fname)
                with open(path, "wb") as wf:
                    wf.write(r.content)
                saved_files.append(path)
            except Exception:
                continue
        data["saved_media_files"] = saved_files

    return data

# Example usage from command line
if __name__ == "__main__":
    import argparse
    import sys
    from textwrap import shorten

    parser = argparse.ArgumentParser(description="Extract public Instagram post/reel metadata using Instaloader")
    parser.add_argument("url", nargs="?", help="Instagram post or reel URL (public). If omitted, you'll be prompted unless --url-file is provided.")
    parser.add_argument("--login-user", help="Instagram username to login (optional)")
    parser.add_argument("--login-pass", help="Instagram password (optional)")
    parser.add_argument("--comments", type=int, default=10, help="How many comments to fetch (sample)")
    parser.add_argument("--save", help="Path to JSON file to save output (optional)")
    parser.add_argument("--access-token", help="Facebook Graph API access token (optional, for insights)")
    parser.add_argument("--save-media", action="store_true", help="Download media URLs to local folder")
    parser.add_argument("--media-dir", default="media", help="Directory to save downloaded media")
    parser.add_argument("--session-file", help="Path to save/load Instagram session (reduces 401/403)")
    parser.add_argument("--proxy", help="HTTP(S) proxy, e.g. http://user:pass@host:port (optional)")
    parser.add_argument("--max-retries", type=int, default=3, help="Max retries on 401/403/429 errors")
    parser.add_argument("--retry-wait", type=int, default=15, help="Base seconds for backoff between retries")
    parser.add_argument("--min-wait", type=float, default=0.0, help="Random wait lower bound between requests (seconds)")
    parser.add_argument("--max-wait", type=float, default=0.0, help="Random wait upper bound between requests (seconds)")
    parser.add_argument("--rpm", type=int, help="Requests per minute cap (simple gate)")
    parser.add_argument("--proxy-pool", help="Path to a file with one proxy per line (rotated per attempt)")
    parser.add_argument("--user-agent", help="Custom User-Agent header")
    parser.add_argument("--user-agent-pool", help="Path to a file with one User-Agent per line (rotated)")
    parser.add_argument("--url-file", help="Path to a text file with one Instagram URL per line")
    # Optional: toggle VPN on rate limit
    parser.add_argument("--vpn-toggle-on-rate-limit", action="store_true", help="On 429/403, cycle WARP via vpn.py then retry")
    parser.add_argument("--vpn-wait", type=float, default=4.0, help="Seconds to wait after reconnecting VPN")
    # Rate-limit-friendly automation
    parser.add_argument("--skip-existing", action="store_true", help="Skip URLs whose shortcode already exists in the Excel output")
    parser.add_argument("--max-per-run", type=int, default=None, help="Process at most N URLs this run (for batching)")
    parser.add_argument("--delay", type=float, default=0.0, help="Base delay in seconds between URL requests (in addition to --rpm)")
    parser.add_argument("--delay-jitter", type=float, default=0.0, help="Add up to this many seconds of random jitter to --delay")
    parser.add_argument("--checkpoint", default=None, help="Path to a checkpoint JSON to resume remaining URLs (optional)")
    parser.add_argument("--ignore-checkpoint", action="store_true", help="Ignore checkpoint file and process all URLs from --url-file")

    args = parser.parse_args()

    # Gather URLs: single or from file
    urls: List[str] = []
    if args.url_file and os.path.exists(args.url_file):
        try:
            with open(args.url_file, "r", encoding="utf-8") as f:
                urls = [ln.strip() for ln in f if ln.strip() and not ln.strip().startswith("#")]
        except Exception as e:
            print(f"Failed to read --url-file: {e}")
            sys.exit(1)
    elif args.url:
        urls = [args.url]
    else:
        try:
            inp = input("Enter public Instagram post/reel URL: ").strip()
            if inp:
                urls = [inp]
        except (KeyboardInterrupt, EOFError):
            urls = []
        if not urls:
            print("\nNo URL provided. Exiting.")
            sys.exit(1)

    # Always define results so downstream code is safe
    results: List[Dict[str, Any]] = []
    total_urls_count = len(urls)
    skipped_count = 0
    error_count = 0
    try:
        # load optional pools
        proxy_pool_list: Optional[List[str]] = None
        ua_pool_list: Optional[List[str]] = None
        if args.proxy_pool and os.path.exists(args.proxy_pool):
            try:
                with open(args.proxy_pool, "r", encoding="utf-8") as pf:
                    proxy_pool_list = [ln.strip() for ln in pf if ln.strip() and not ln.strip().startswith("#")]
            except Exception:
                proxy_pool_list = None
        if args.user_agent_pool and os.path.exists(args.user_agent_pool):
            try:
                with open(args.user_agent_pool, "r", encoding="utf-8") as uf:
                    ua_pool_list = [ln.strip() for ln in uf if ln.strip() and not ln.strip().startswith("#")]
            except Exception:
                ua_pool_list = None

        # Optional: build processed set from existing Excel to skip duplicates quickly
        processed_shortcodes: set = set()
        processed_urls: set = set()  # Also track full URLs for matching
        excel_path_tmp = os.path.join(os.path.dirname(__file__), 'instagram_post_data.xlsx')
        if args.skip_existing and os.path.exists(excel_path_tmp):
            try:
                import pandas as pd
                existing_df_tmp = pd.read_excel(excel_path_tmp, usecols=lambda c: True)
                
                # Extract shortcodes from 'shortcode' column if it exists
                if 'shortcode' in existing_df_tmp.columns:
                    for s in existing_df_tmp['shortcode'].dropna().astype(str):
                        processed_shortcodes.add(str(s).strip())
                
                # Extract shortcodes and URLs from 'url' column if it exists
                if 'url' in existing_df_tmp.columns:
                    for u in existing_df_tmp['url'].dropna().astype(str):
                        u_str = str(u).strip()
                        processed_urls.add(u_str)
                        # Also extract shortcode from URL and add it
                        sc_from_url = shortcode_from_url(u_str)
                        if sc_from_url:
                            processed_shortcodes.add(sc_from_url)
                        # Also add normalized URL (remove query params for matching)
                        normalized_url = u_str.split('?')[0].split('#')[0].rstrip('/')
                        processed_urls.add(normalized_url)
            except Exception as e:
                print(f"[Warning] Could not read Excel for skip-existing check: {e}")
                processed_shortcodes = set()
                processed_urls = set()

        # Optional: load checkpoint to resume
        checkpoint_path = args.checkpoint
        if checkpoint_path and os.path.exists(checkpoint_path) and not args.ignore_checkpoint:
            try:
                with open(checkpoint_path, 'r', encoding='utf-8') as cf:
                    pending_from_ckpt = json.load(cf)
                if isinstance(pending_from_ckpt, list) and pending_from_ckpt:
                    original_count = len(urls)
                    urls = [u for u in pending_from_ckpt if isinstance(u, str) and u.strip()]
                    print(f"[Checkpoint] Loaded {len(urls)} URL(s) from checkpoint (original: {original_count})")
                    print(f"[Checkpoint] To process all URLs from file, use --ignore-checkpoint")
            except Exception as e:
                print(f"[Warning] Could not load checkpoint: {e}")
                pass
        elif checkpoint_path and os.path.exists(checkpoint_path) and args.ignore_checkpoint:
            print(f"[Checkpoint] Ignoring checkpoint file (--ignore-checkpoint used)")

        # Throttle helpers
        import time
        last_request_ts_holder = {"ts": None}  # mutable holder so inner function can update it

        def sleep_between_urls():
            # Respect RPM gate if provided in inner calls too, but add explicit inter-URL sleep
            base_delay = max(0.0, float(args.delay or 0.0))
            jitter = max(0.0, float(args.delay_jitter or 0.0))
            extra = random.uniform(0.0, jitter) if jitter > 0 else 0.0
            # Token spacing from RPM: if set, keep at least 60/rpm seconds between URLs
            minimal = 0.0
            if args.rpm and args.rpm > 0:
                minimal = max(minimal, 60.0 / float(args.rpm))
            target_gap = max(minimal, base_delay + extra)
            if last_request_ts_holder["ts"] is not None and target_gap > 0:
                elapsed = time.time() - last_request_ts_holder["ts"]
                remain = target_gap - elapsed
                if remain > 0:
                    try:
                        time.sleep(remain)
                    except Exception:
                        pass
            last_request_ts_holder["ts"] = time.time()

        # Backoff helper for 429/403
        def backoff_sleep(attempt: int, base: int) -> None:
            try:
                sleep_s = max(base, 10) * (2 ** min(attempt, 5))
                sleep_s = min(sleep_s, 15 * 60)
                time.sleep(sleep_s)
            except Exception:
                pass

        # process one or many URLs with rate-limit-aware loop
        processed_count = 0
        pending_urls = list(urls)
        
        print(f"\n--- Processing {len(pending_urls)} URL(s) ---")
        if args.skip_existing:
            print(f"Found {len(processed_shortcodes)} already processed shortcode(s) and {len(processed_urls)} URL(s) in Excel. Skipping duplicates.")
            # Show sample of what's in Excel for debugging
            if len(processed_shortcodes) > 0:
                sample_sc = list(processed_shortcodes)[:3]
                print(f"  Sample shortcodes from Excel: {sample_sc}")

        for idx, url in enumerate(pending_urls):
            if args.max_per_run is not None and processed_count >= args.max_per_run:
                break

            # Check if URL should be skipped (by shortcode or normalized URL)
            should_skip = False
            skip_reason = ""
            if args.skip_existing:
                sc = shortcode_from_url(url)
                normalized_url = url.split('?')[0].split('#')[0].rstrip('/')
                
                # Check if shortcode matches
                if sc and sc in processed_shortcodes:
                    should_skip = True
                    skip_reason = f"shortcode '{sc}'"
                # Check if normalized URL matches
                elif normalized_url in processed_urls:
                    should_skip = True
                    skip_reason = f"normalized URL '{normalized_url}'"
                # Also check if full URL matches (with or without query params)
                elif url in processed_urls:
                    should_skip = True
                    skip_reason = "full URL"
                
                if should_skip:
                    skipped_count += 1
                    if skipped_count <= 5 or (skipped_count % 10 == 0):
                        print(f"[{idx+1}/{len(pending_urls)}] Skipping {url} (matched by {skip_reason})")
                    continue

            print(f"[{idx+1}/{len(pending_urls)}] Processing {url}...")
            sleep_between_urls()

            # Retry with explicit handling of rate-limit errors
            rl_attempt = 0
            while True:
                try:
                    # Rotate proxy and user-agent per URL to distribute load
                    sel_proxy_for_url = None
                    if proxy_pool_list and len(proxy_pool_list) > 0:
                        sel_proxy_for_url = proxy_pool_list[idx % len(proxy_pool_list)]
                    sel_ua_for_url = None
                    if ua_pool_list and len(ua_pool_list) > 0:
                        try:
                            sel_ua_for_url = random.choice(ua_pool_list)
                        except Exception:
                            sel_ua_for_url = None
                    data = get_full_instagram_post_data(
                        url,
                        login_user=args.login_user,
                        login_pass=args.login_pass,
                        fetch_comments=args.comments,
                        access_token=args.access_token,
                        save_media=args.save_media,
                        media_dir=args.media_dir,
                        session_file=args.session_file,
                        proxy=sel_proxy_for_url or args.proxy,
                        max_retries=args.max_retries,
                        retry_wait=args.retry_wait,
                        min_wait=args.min_wait,
                        max_wait=args.max_wait,
                        requests_per_minute=args.rpm,
                        proxy_pool=proxy_pool_list,
                        user_agent=sel_ua_for_url or args.user_agent,
                        user_agent_pool=ua_pool_list,
                    )
                    # add scraped timestamp (UTC ISO) for Excel export and records
                    try:
                        data['scraped_at_utc'] = datetime.utcnow().isoformat()
                    except Exception:
                        data['scraped_at_utc'] = None
                    results.append(data)
                    processed_count += 1
                    print(f"  ✓ Success! (shortcode: {data.get('shortcode', 'N/A')})")
                    if args.skip_existing:
                        if data.get('shortcode'):
                            processed_shortcodes.add(str(data['shortcode']))
                        elif data.get('url'):
                            processed_shortcodes.add(str(data['url']))
                    break
                except Exception as e:
                    emsg = str(e).lower()
                    if any(tok in emsg for tok in ["429", "too many requests", "rate limit", "http error code 429"]):
                        # Optional VPN cycle before backoff
                        if args.vpn_toggle_on_rate_limit and vpn is not None:
                            try:
                                print(f"  ⚠ Rate limit detected. Cycling VPN...")
                                vpn.warp_cycle(disconnect_wait=max(1.0, (args.retry_wait or 1) / 3), reconnect_wait=max(1.0, args.vpn_wait or 1.0))
                            except Exception:
                                pass
                        rl_attempt += 1
                        if args.max_retries > 0 and rl_attempt <= args.max_retries:
                            print(f"  ⚠ Rate limited (attempt {rl_attempt}/{args.max_retries}). Backing off...")
                            backoff_sleep(rl_attempt, base=max(30, args.retry_wait or 15))
                            continue
                        else:
                            print(f"  ✗ Rate limit error (max retries reached or disabled): {e}")
                            error_count += 1
                            break
                    if any(tok in emsg for tok in ["403", "forbidden", "login required"]):
                        if args.vpn_toggle_on_rate_limit and vpn is not None:
                            try:
                                print(f"  ⚠ 403 Forbidden. Cycling VPN...")
                                vpn.warp_cycle(disconnect_wait=max(1.0, (args.retry_wait or 1) / 3), reconnect_wait=max(1.0, args.vpn_wait or 1.0))
                            except Exception:
                                pass
                        rl_attempt += 1
                        if args.max_retries > 0 and rl_attempt <= args.max_retries:
                            print(f"  ⚠ 403 Forbidden (attempt {rl_attempt}/{args.max_retries}). Backing off...")
                            backoff_sleep(rl_attempt, base=max(60, args.retry_wait or 15))
                            continue
                        else:
                            print(f"  ✗ 403 Forbidden error (max retries reached or disabled): {e}")
                            error_count += 1
                            break
                    print(f"  ✗ Error on URL {url}: {e}")
                    error_count += 1
                    break

            # Save checkpoint of remaining URLs if requested
            if checkpoint_path:
                try:
                    remaining = pending_urls[idx + 1:]
                    if args.max_per_run is not None and processed_count >= args.max_per_run:
                        remaining = pending_urls[idx:]
                    with open(checkpoint_path, 'w', encoding='utf-8') as cf:
                        json.dump(remaining, cf, indent=2)
                except Exception:
                    pass
    except Exception as e:
        print(f"\nFatal Error: {e}")
        import traceback
        traceback.print_exc()
        # fall through to handle empty results

    # --- EXPORT TO EXCEL (append rows across runs) ---
    try:
        import pandas as pd
        excel_path = os.path.join(os.path.dirname(__file__), 'instagram_post_data.xlsx')

        def to_row(d: Dict[str, Any]) -> Dict[str, Any]:
            flat = d.copy()
            seo = flat.pop('seo', {}) or {}
            for k, v in seo.items():
                flat[f'seo_{k}'] = v
            if isinstance(flat.get('hashtags'), list):
                flat['hashtags'] = ', '.join(flat['hashtags'])
            if isinstance(flat.get('comments_sample'), list):
                flat['comments_sample_count'] = len(flat['comments_sample'])
            return flat

        if not results:
            raise RuntimeError("No successful results to export.")
        rows: List[Dict[str, Any]] = [to_row(d) for d in results]

        new_df = pd.DataFrame(rows)

        # Prefer true append without re-reading entire file to avoid overwriting
        if os.path.exists(excel_path):
            try:
                from openpyxl import load_workbook  # engine used by pandas for Excel IO
                book = load_workbook(excel_path)
                sheet_name = book.sheetnames[0] if book.sheetnames else 'Sheet1'
                start_row = book[sheet_name].max_row  # append after last non-empty row

                with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    new_df.to_excel(writer, index=False, header=False, startrow=start_row, sheet_name=sheet_name)
            except Exception:
                # Fallback: concatenate in-memory (previous behavior)
                try:
                    existing_df = pd.read_excel(excel_path)
                except Exception:
                    existing_df = pd.DataFrame()
                combined_df = pd.concat([existing_df, new_df], ignore_index=True, sort=False)
                # De-duplicate by shortcode/url if present
                dedupe_key = 'shortcode' if 'shortcode' in combined_df.columns else ('url' if 'url' in combined_df.columns else None)
                if dedupe_key:
                    combined_df.drop_duplicates(subset=[dedupe_key], keep='last', inplace=True)
                combined_df.to_excel(excel_path, index=False)
        else:
            # First time: create file with headers
            new_df.to_excel(excel_path, index=False)

        print(f"\n--- Data exported to {excel_path} ---")
    except ImportError:
        print("[Warning] pandas not installed. Skipping Excel export.")
    except Exception as e:
        print(f"[Warning] Failed to write Excel file: {e}")

    # Print JSON to console
    print("\n--- Full JSON Output ---")
    if results:
        print(json.dumps(results, indent=2, ensure_ascii=False))
    else:
        print("[]")

    # Human-readable summary
    print("\n--- Summary ---")
    print(f"Total URLs: {total_urls_count}")
    print(f"Successfully processed: {len(results)}")
    print(f"Errors: {error_count}")
    print(f"Skipped (--skip-existing): {skipped_count}")
    
    if results:
        print("\nProcessed URLs:")
        for d in results:
            try:
                print(f"- {d.get('url')} (score={d.get('seo',{}).get('seo_score')})")
            except Exception:
                print(f"- {d.get('url')}")
    elif args.skip_existing:
        print("\n⚠ All URLs were skipped because they already exist in Excel.")
        print("   Remove --skip-existing to reprocess them.")

    # Optionally save JSON to file
    if args.save and results:
        try:
            with open(args.save, "w", encoding="utf-8") as f:
                json.dump(results, f, indent=2, ensure_ascii=False)
            print(f"\nSaved JSON to {args.save}")
        except Exception as e:
            print(f"Failed to save JSON: {e}")